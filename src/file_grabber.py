#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
文件抓取器 - 抓取文件夾內的文件名並創建超鏈接列表

作者: David-CB666
版本: v2.1
日期: 2026-06-02
"""

import os
import sys
import io
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
except ImportError:
    print("[ERROR] 請安裝 openpyxl: pip install openpyxl")
    sys.exit(1)


class FileGrabber:
    """
    文件抓取器
    
    功能：
    1. 掃描文件夾內的文件
    2. 創建超鏈接列表
    3. 支持擴展名過濾
    
    使用示例：
        grabber = FileGrabber()
        files = grabber.grab_files("./files", extensions=[".pdf"])
        grabber.write_to_worksheet(files, "output.xlsx", "Sheet1")
    """
    
    def __init__(self):
        """初始化文件抓取器"""
        pass
    
    def grab_files(
        self,
        folder: str,
        include_subfolders: bool = True,
        extensions: Optional[List[str]] = None
    ) -> List[Dict[str, str]]:
        """
        抓取文件列表
        
        Args:
            folder: 目標文件夾
            include_subfolders: 是否包含子文件夾
            extensions: 擴展名過濾（None = 全部）
        
        Returns:
            [{"path": "完整路徑", "name": "文件名", "ext": "擴展名"}, ...]
        """
        folder_path = Path(folder)
        if not folder_path.exists():
            raise FileNotFoundError(f"文件夾不存在: {folder}")
        
        # 處理擴展名
        if extensions:
            extensions = [ext.lower() for ext in extensions]
            extensions = ['.' + ext if not ext.startswith('.') else ext for ext in extensions]
        
        files = []
        
        # 遞歸掃描
        if include_subfolders:
            pattern = "**/*"
        else:
            pattern = "*"
        
        for file_path in folder_path.glob(pattern):
            if file_path.is_file():
                # 擴展名過濾
                if extensions:
                    if file_path.suffix.lower() not in extensions:
                        continue
                
                files.append({
                    "path": str(file_path.absolute()),
                    "name": file_path.name,
                    "ext": file_path.suffix,
                    "size": file_path.stat().st_size,
                    "modified": datetime.fromtimestamp(file_path.stat().st_mtime)
                })
        
        print(f"[OK] 已抓取 {len(files)} 個文件")
        return files
    
    def write_to_worksheet(
        self,
        files: List[Dict[str, str]],
        workbook_path: str,
        sheet_name: str = "Sheet1",
        start_cell: str = "A1",
        create_hyperlinks: bool = True
    ) -> str:
        """
        寫入工作表
        
        Args:
            files: 文件列表（從 grab_files 返回）
            workbook_path: 工作簿路徑
            sheet_name: 工作表名稱
            start_cell: 起始單元格
            create_hyperlinks: 是否創建超鏈接
        
        Returns:
            工作簿路徑
        """
        wb_path = Path(workbook_path)
        
        # 載入或創建工作簿
        if wb_path.exists():
            wb = load_workbook(wb_path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
        
        # 解析起始單元格
        import re
        match = re.match(r'([A-Z]+)(\d+)', start_cell.upper())
        if match:
            start_col = ord(match.group(1)) - ord('A') + 1
            start_row = int(match.group(2))
        else:
            start_col, start_row = 1, 1
        
        # 寫入數據
        for idx, file_info in enumerate(files):
            row = start_row + idx
            cell = ws.cell(row=row, column=start_col)
            
            if create_hyperlinks:
                # 創建超鏈接
                cell.hyperlink = file_info["path"]
                cell.value = file_info["name"]
                cell.font = Font(color="0000FF", underline="single")
            else:
                cell.value = file_info["name"]
        
        # 保存
        wb.save(wb_path)
        print(f"[OK] 已保存: {wb_path}")
        
        return str(wb_path)


def main():
    """命令行入口"""
    import argparse
    
    parser = argparse.ArgumentParser(description="文件抓取器")
    parser.add_argument("--folder", required=True, help="目標文件夾")
    parser.add_argument("--output", required=True, help="輸出工作簿路徑")
    parser.add_argument("--sheet", default="Sheet1", help="工作表名稱")
    parser.add_argument("--start", default="A1", help="起始單元格")
    parser.add_argument("--ext", default=None, help="擴展名過濾（逗號分隔）")
    parser.add_argument("--no-subfolders", action="store_true", help="不包含子文件夾")
    parser.add_argument("--no-links", action="store_true", help="不創建超鏈接")
    
    args = parser.parse_args()
    
    extensions = None
    if args.ext:
        extensions = [ext.strip() for ext in args.ext.split(',')]
    
    grabber = FileGrabber()
    
    files = grabber.grab_files(
        folder=args.folder,
        include_subfolders=not args.no_subfolders,
        extensions=extensions
    )
    
    grabber.write_to_worksheet(
        files=files,
        workbook_path=args.output,
        sheet_name=args.sheet,
        start_cell=args.start,
        create_hyperlinks=not args.no_links
    )


if __name__ == "__main__":
    main()
