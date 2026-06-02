#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BQ 頁合併器 - 將申請表 PDF 與 BQ 頁面合併

作者: David-CB666
版本: v2.1
日期: 2026-06-02
"""

import os
import re
import sys
import io
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    import fitz  # PyMuPDF
except ImportError:
    print("[ERROR] 請安裝 PyMuPDF: pip install PyMuPDF")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("[ERROR] 請安裝 openpyxl: pip install openpyxl")
    sys.exit(1)


class BQMerger:
    """
    BQ 頁合併器
    
    功能：
    1. 讀取總表獲取 EL 編號 + 材料名 + BQ 編號
    2. 在 BQ PDF 中搜索每個編號所在的頁碼（正則匹配）
    3. 將報批表 PDF 與對應 BQ 整頁合併
    4. 輸出文件名格式：EL-XXX 材料名.pdf
    
    使用示例：
        merger = BQMerger()
        merger.load_bq_pdf("BQ.pdf")
        merger.load_zongbiao("總表.xlsx")
        merger.merge_pdfs("./input", "./output")
    """
    
    def __init__(self):
        """初始化 BQ 合併器"""
        self.bq_doc: Optional[fitz.Document] = None
        self.bq_page_index: Dict[str, int] = {}  # {BQ編號: 頁碼}
        self.zongbiao_data: List[Dict[str, str]] = []
    
    def load_bq_pdf(self, bq_path: str) -> bool:
        """
        載入 BQ PDF 並建立頁碼索引
        
        Args:
            bq_path: BQ PDF 路徑
        
        Returns:
            是否成功載入
        """
        try:
            self.bq_doc = fitz.open(bq_path)
            
            print(f"[OK] 已載入 BQ PDF: {bq_path}")
            print(f"  總頁數: {self.bq_doc.page_count}")
            
            # 建立頁碼索引
            self._build_page_index()
            
            return True
            
        except Exception as e:
            print(f"[ERROR] 載入 BQ PDF 失敗: {e}")
            return False
    
    def _build_page_index(self):
        """建立 BQ 編號 → 頁碼索引"""
        self.bq_page_index.clear()
        
        for page_num in range(self.bq_doc.page_count):
            text = self.bq_doc[page_num].get_text()
            
            # 匹配 BQ 編號格式：1.1, 2.1.3, 1.1-a 等
            matches = re.findall(r'\b(\d+\.\d+(?:[-.]\d+)?)\b', text)
            
            for m in matches:
                if m not in self.bq_page_index:
                    self.bq_page_index[m] = page_num + 1  # 轉為 1-based 頁碼
        
        print(f"[OK] 已建立頁碼索引: {len(self.bq_page_index)} 個 BQ 編號")
    
    def load_zongbiao(
        self,
        zongbiao_path: str,
        sheet_index: int = 0,
        start_row: int = 7,
        col_bq: int = 1,
        col_el: int = 2,
        col_name: int = 3
    ) -> int:
        """
        讀取數據源中的條目列表
        
        Args:
            zongbiao_path: 總表 Excel 路徑
            sheet_index: 工作表索引
            start_row: 數據起始行
            col_bq: BQ 編號列（1-based）
            col_el: EL 編號列
            col_name: 材料名列
        
        Returns:
            讀取的材料數量
        """
        try:
            wb = openpyxl.load_workbook(zongbiao_path, data_only=True)
            ws = wb.worksheets[sheet_index]
            
            self.zongbiao_data.clear()
            
            for row_idx in range(start_row, ws.max_row + 1):
                bq = ws.cell(row_idx, col_bq).value
                el = ws.cell(row_idx, col_el).value
                name = ws.cell(row_idx, col_name).value
                
                # 跳過空行
                if bq is None and el is None:
                    break
                
                self.zongbiao_data.append({
                    'bq': str(bq).strip() if bq else '',
                    'el': str(el).strip() if el else '',
                    'name': str(name).strip() if name else ''
                })
            
            wb.close()
            
            print(f"[OK] 已讀取總表: {zongbiao_path}")
            print(f"  材料數: {len(self.zongbiao_data)}")
            
            return len(self.zongbiao_data)
            
        except Exception as e:
            print(f"[ERROR] 讀取總表失敗: {e}")
            return 0
    
    def match_bq_pages(self) -> Tuple[int, int]:
        """
        匹配 BQ 頁碼
        
        Returns:
            (成功數, 失敗數)
        """
        if not self.bq_doc:
            raise RuntimeError("請先載入 BQ PDF")
        
        if not self.zongbiao_data:
            raise RuntimeError("請先載入總表")
        
        success = 0
        failed = 0
        
        for item in self.zongbiao_data:
            bq = item['bq']
            
            if bq in self.bq_page_index:
                # 直接匹配
                item['bq_page'] = self.bq_page_index[bq]
                success += 1
            else:
                # 嘗試模糊匹配（去掉後綴）
                found = False
                parts = bq.rsplit('.', 1)
                
                for n in range(len(parts), 0, -1):
                    base = '.'.join(parts[:n])
                    if base in self.bq_page_index:
                        item['bq_page'] = self.bq_page_index[base]
                        found = True
                        success += 1
                        break
                
                if not found:
                    item['bq_page'] = None
                    failed += 1
                    print(f"  [WARN] 未找到 BQ 頁: {item['el']}  BQ={bq}")
        
        print(f"[OK] BQ 頁匹配完成: {success} 成功, {failed} 失敗")
        
        return success, failed
    
    def merge_pdfs(
        self,
        input_dir: str,
        output_dir: str
    ) -> Tuple[int, int]:
        """
        合併 PDF
        
        Args:
            input_dir: 報批表 PDF 輸入目錄
            output_dir: 合併後 PDF 輸出目錄
        
        Returns:
            (成功數, 跳過數)
        """
        if not self.bq_doc:
            raise RuntimeError("請先載入 BQ PDF")
        
        # 創建輸出目錄
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        input_path = Path(input_dir)
        
        success = 0
        skipped = 0
        
        print(f"\n[INFO] 開始合併 PDF...")
        
        for item in self.zongbiao_data:
            el = item['el']
            bq_page = item.get('bq_page')
            
            # 查找報批表 PDF
            baopiao_file = None
            for f in input_path.iterdir():
                if f.suffix.lower() == '.pdf' and f.name.lower().startswith(el.lower()):
                    baopiao_file = f
                    break
            
            if not baopiao_file:
                print(f"  [SKIP] 無報批表 PDF: {el}")
                skipped += 1
                continue
            
            if bq_page is None:
                print(f"  [SKIP] 無 BQ 頁: {el}")
                skipped += 1
                continue
            
            # 合併
            try:
                output = fitz.open()
                
                # 插入報批表
                baopiao_doc = fitz.open(str(baopiao_file))
                output.insert_pdf(baopiao_doc)
                baopiao_doc.close()
                
                # 插入 BQ 頁
                output.insert_pdf(
                    self.bq_doc,
                    from_page=bq_page - 1,
                    to_page=bq_page - 1
                )
                
                # 生成輸出文件名
                safe_name = re.sub(r'[\\/:*?"<>|]', '-', item['name'])
                output_name = f"{el} {safe_name}.pdf"
                output_file = output_path / output_name
                
                output.save(str(output_file))
                output.close()
                
                print(f"  [OK] {el}  {item['name'][:18]}  → BQ p{bq_page}")
                success += 1
                
            except Exception as e:
                print(f"  [ERROR] 合併失敗: {el}  {e}")
                skipped += 1
        
        print(f"\n[OK] 合併完成: {success} 成功, {skipped} 跳過")
        print(f"  輸出目錄: {output_dir}")
        
        return success, skipped
    
    def close(self):
        """關閉 BQ PDF"""
        if self.bq_doc:
            self.bq_doc.close()
            self.bq_doc = None


def main():
    """命令行入口"""
    import argparse
    
    parser = argparse.ArgumentParser(description="BQ 頁合併器")
    parser.add_argument("--zongbiao", required=True, help="總表 Excel 路徑")
    parser.add_argument("--bq", required=True, help="BQ PDF 路徑")
    parser.add_argument("--input", required=True, help="報批表 PDF 輸入目錄")
    parser.add_argument("--output", required=True, help="合併後 PDF 輸出目錄")
    parser.add_argument("--sheet", type=int, default=0, help="工作表索引（默認 0）")
    parser.add_argument("--start-row", type=int, default=7, help="數據起始行（默認 7）")
    
    args = parser.parse_args()
    
    merger = BQMerger()
    
    # 載入 BQ PDF
    if not merger.load_bq_pdf(args.bq):
        sys.exit(1)
    
    # 載入總表
    if not merger.load_zongbiao(args.zongbiao, args.sheet, args.start_row):
        sys.exit(1)
    
    # 匹配 BQ 頁
    merger.match_bq_pages()
    
    # 合併 PDF
    merger.merge_pdfs(args.input, args.output)
    
    merger.close()
    
    print("\n✅ 完成！")


if __name__ == "__main__":
    main()
