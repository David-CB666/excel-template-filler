#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
自動鏈接器 - 根據單元格內容自動匹配文件並創建超鏈接

作者: David-CB666
版本: v2.1
日期: 2026-06-02
"""

import os
import re
import sys
import io
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
except ImportError:
    print("[ERROR] 請安裝 openpyxl: pip install openpyxl")
    sys.exit(1)


class AutoLinker:
    """
    自動鏈接器
    
    功能：
    1. 根據單元格內容搜索文件
    2. 創建超鏈接
    3. 支持通配符匹配
    
    使用示例：
        linker = AutoLinker("summary.xlsx", 1, "./files")
        result = linker.convert_to_hyperlinks()
        print(f"成功: {result['success']}, 跳過: {result['skipped']}")
    """
    
    def __init__(
        self,
        workbook: str,
        target_column: int = 1,
        search_root: str = "."
    ):
        """
        初始化自動鏈接器
        
        Args:
            workbook: 工作簿路徑
            target_column: 目標列（1 = A 列）
            search_root: 搜索根目錄
        """
        self.workbook_path = Path(workbook)
        self.target_column = target_column
        self.search_root = Path(search_root)
        
        # 驗證
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"工作簿不存在: {workbook}")
        if not self.search_root.exists():
            raise FileNotFoundError(f"搜索目錄不存在: {search_root}")
        
        # 載入工作簿
        self.wb = load_workbook(self.workbook_path)
        self.ws = self.wb.active
        
        # 文件索引
        self.file_index: Dict[str, Path] = {}  # {文件名（不含擴展名）: 完整路徑}
        
    def build_file_index(
        self,
        extensions: List[str] = None,
        include_subfolders: bool = True
    ):
        """
        構建文件索引
        
        Args:
            extensions: 擴展名列表（如 [".pdf", ".docx"]）
            include_subfolders: 是否包含子文件夾
        """
        self.file_index.clear()
        
        if extensions:
            extensions = [ext.lower() for ext in extensions]
            if not all(ext.startswith('.') for ext in extensions):
                extensions = ['.' + ext if not ext.startswith('.') else ext for ext in extensions]
        
        if include_subfolders:
            pattern = "**/*"
        else:
            pattern = "*"
        
        for file_path in self.search_root.glob(pattern):
            if file_path.is_file():
                if extensions:
                    if file_path.suffix.lower() not in extensions:
                        continue
                
                # 索引鍵：文件名（不含擴展名）
                key = file_path.stem.lower()
                self.file_index[key] = file_path
        
        print(f"[OK] 已索引 {len(self.file_index)} 個文件")
    
    def find_file(
        self,
        cell_value: str,
        use_wildcard: bool = True
    ) -> Optional[Path]:
        """
        查找匹配的文件
        
        Args:
            cell_value: 單元格值
            use_wildcard: 是否使用通配符匹配
        
        Returns:
            匹配的文件路徑，未找到返回 None
        """
        if not cell_value:
            return None
        
        cell_value_lower = cell_value.lower().strip()
        
        # 1. 精確匹配
        if cell_value_lower in self.file_index:
            return self.file_index[cell_value_lower]
        
        # 2. 通配符匹配
        if use_wildcard:
            for key, path in self.file_index.items():
                if cell_value_lower in key:
                    return path
        
        return None
    
    def convert_to_hyperlinks(
        self,
        extensions: List[str] = None,
        include_subfolders: bool = True,
        use_wildcard: bool = True,
        start_row: int = 2
    ) -> Dict[str, int]:
        """
        轉換單元格為超鏈接
        
        Args:
            extensions: 文件擴展名列表
            include_subfolders: 是否包含子文件夾
            use_wildcard: 是否使用通配符匹配
            start_row: 開始行（默認 2，跳過表頭）
        
        Returns:
            {"success": 成功數, "skipped": 跳過數}
        """
        # 構建文件索引
        self.build_file_index(extensions, include_subfolders)
        
        success_count = 0
        skipped_count = 0
        
        # 遍歷目標列
        for row_idx in range(start_row, self.ws.max_row + 1):
            cell = self.ws.cell(row=row_idx, column=self.target_column)
            cell_value = str(cell.value).strip() if cell.value else ""
            
            if not cell_value:
                skipped_count += 1
                continue
            
            # 查找文件
            matched_file = self.find_file(cell_value, use_wildcard)
            
            if matched_file:
                # 計算相對路徑
                try:
                    rel_path = matched_file.relative_to(self.workbook_path.parent)
                    hyperlink_path = f"./{rel_path}"
                except ValueError:
                    hyperlink_path = str(matched_file)
                
                # 清空單元格
                cell.value = None
                
                # 創建超鏈接
                cell.hyperlink = hyperlink_path
                cell.value = cell_value  # 顯示文本
                
                # 設置樣式
                cell.font = Font(color="0000FF", underline="single")
                
                success_count += 1
                print(f"[OK] 第 {row_idx} 行: {cell_value} -> {matched_file.name}")
            else:
                skipped_count += 1
                print(f"[SKIP] 第 {row_idx} 行: {cell_value} 未找到匹配文件")
        
        # 保存
        self.wb.save(self.workbook_path)
        print(f"\n[OK] 已保存: {self.workbook_path}")
        
        return {"success": success_count, "skipped": skipped_count}
    
    def close(self):
        """關閉工作簿"""
        self.wb.close()


def main():
    """命令行入口"""
    import argparse
    
    parser = argparse.ArgumentParser(description="自動鏈接器")
    parser.add_argument("--workbook", required=True, help="工作簿路徑")
    parser.add_argument("--column", type=int, default=1, help="目標列（默認 1）")
    parser.add_argument("--root", default=".", help="搜索根目錄")
    parser.add_argument("--ext", default=".pdf,.docx,.xlsx", help="擴展名（逗號分隔）")
    parser.add_argument("--no-subfolders", action="store_true", help="不包含子文件夾")
    parser.add_argument("--exact", action="store_true", help="精確匹配（不用通配符）")
    
    args = parser.parse_args()
    
    extensions = [ext.strip() for ext in args.ext.split(',')]
    
    linker = AutoLinker(
        workbook=args.workbook,
        target_column=args.column,
        search_root=args.root
    )
    
    result = linker.convert_to_hyperlinks(
        extensions=extensions,
        include_subfolders=not args.no_subfolders,
        use_wildcard=not args.exact
    )
    
    print(f"\n結果: 成功 {result['success']}, 跳過 {result['skipped']}")
    
    linker.close()


if __name__ == "__main__":
    main()
