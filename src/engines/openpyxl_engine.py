#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
openpyxl 引擎 - 使用 openpyxl 處理無圖片模板

適用於：無圖片、無打印設置的簡單模板

作者: David-CB666
版本: v2.1
日期: 2026-06-02
"""

import os
import re
import sys
import io
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("[ERROR] 請安裝 openpyxl: pip install openpyxl")
    sys.exit(1)

from .base_engine import BaseEngine


class OpenPYXLEngine(BaseEngine):
    """
    openpyxl 引擎
    
    適用於無圖片的簡單模板
    優點：代碼簡潔、易於維護、純 Python
    缺點：會丟失圖片和打印設置
    """
    
    def __init__(self):
        """初始化 openpyxl 引擎"""
        super().__init__()
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None
        self.placeholders: Dict[str, List[str]] = {}  # {字段名: [單元格地址列表]}
        self.header_map: Dict[str, int] = {}  # {表頭名: 列號}
    
    def load_template(self, template_path: str) -> bool:
        """
        載入模板文件
        
        Args:
            template_path: 模板 Excel 路徑
        
        Returns:
            是否成功載入
        """
        try:
            self.template_path = template_path
            self.workbook = load_workbook(template_path)
            self.worksheet = self.workbook.active
            self.is_loaded = True
            
            print(f"[OK] 已載入模板: {template_path}")
            return True
            
        except Exception as e:
            print(f"[ERROR] 載入模板失敗: {e}")
            return False
    
    def scan_placeholders(self) -> List[str]:
        """
        掃描模板中的佔位符
        
        Returns:
            佔位符列表
        """
        if not self.worksheet:
            raise RuntimeError("請先載入模板")
        
        self.placeholders.clear()
        
        # 正則表達式匹配 {{字段名}} 或 {字段名}
        double_brace_pattern = re.compile(r'\{\{([^}]+)\}\}')
        single_brace_pattern = re.compile(r'\{([^}]+)\}')
        
        for row in self.worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value = cell.value.strip()
                    
                    # 優先匹配雙花括號
                    match = double_brace_pattern.match(value)
                    if match:
                        placeholder_name = match.group(1).strip()
                    else:
                        # 嘗試匹配單花括號
                        match = single_brace_pattern.match(value)
                        if match:
                            placeholder_name = match.group(1).strip()
                        else:
                            continue
                    
                    # 記錄佔位符位置
                    if placeholder_name not in self.placeholders:
                        self.placeholders[placeholder_name] = []
                    self.placeholders[placeholder_name].append(cell.coordinate)
        
        return list(self.placeholders.keys())
    
    def scan_headers(self) -> Dict[str, int]:
        """
        掃描數據源表頭（第一行）
        
        Returns:
            {表頭名: 列號}
        """
        self.header_map.clear()
        
        if not self.worksheet:
            raise RuntimeError("請先載入模板")
        
        for col_idx, cell in enumerate(self.worksheet[1], start=1):
            if cell.value:
                header_name = str(cell.value).strip()
                self.header_map[header_name] = col_idx
        
        return self.header_map
    
    def fill_template(
        self,
        data: Dict[str, Any],
        field_map: Dict[str, str]
    ) -> Worksheet:
        """
        填充單個模板
        
        Args:
            data: 數據字典
            field_map: 字段映射 {佔位符: 字段名}
        
        Returns:
            填充後的工作表
        """
        if not self.worksheet:
            raise RuntimeError("請先載入模板")
        
        # 替換佔位符
        for placeholder, field_name in field_map.items():
            if placeholder in self.placeholders:
                cell_value = data.get(field_name, "")
                
                # 處理空值
                if cell_value is None or cell_value == "":
                    cell_value = ""
                
                # 替換所有出現的佔位符
                for addr in self.placeholders[placeholder]:
                    self.worksheet[addr].value = cell_value
        
        return self.worksheet
    
    def fill_and_export(
        self,
        data_list: List[Dict[str, Any]],
        field_map: Dict[str, str],
        output_path: str
    ) -> str:
        """
        批量填充並導出
        
        Args:
            data_list: 數據列表
            field_map: 字段映射 {佔位符: 字段名}
            output_path: 輸出路徑
        
        Returns:
            輸出文件路徑
        """
        if not self.workbook:
            raise RuntimeError("請先載入模板")
        
        print(f"[INFO] 開始生成 {len(data_list)} 個 Sheet...")
        
        # 創建新工作簿
        merged_wb = Workbook()
        merged_wb.remove(merged_wb.active)  # 刪除默認工作表
        
        for i, data in enumerate(data_list):
            # 創建模板副本
            new_ws = self.worksheet
            
            # 填充數據
            for placeholder, field_name in field_map.items():
                if placeholder in self.placeholders:
                    cell_value = data.get(field_name, "")
                    
                    if cell_value is None or cell_value == "":
                        cell_value = ""
                    
                    for addr in self.placeholders[placeholder]:
                        new_ws[addr].value = cell_value
            
            # 生成工作表名稱
            first_field = list(field_map.values())[0] if field_map else None
            sheet_name = str(data.get(first_field, f"Sheet{i + 1}"))[:31] if first_field else f"Sheet{i + 1}"
            
            # 清理工作表名稱
            sheet_name = self._clean_sheet_name(sheet_name)
            
            # 複製到合併工作簿
            new_ws.title = sheet_name
            copied_ws = merged_wb.create_sheet(title=sheet_name)
            
            # 複製內容
            for row in new_ws.iter_rows():
                for cell in row:
                    new_cell = copied_ws[cell.coordinate]
                    new_cell.value = cell.value
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.alignment = cell.alignment.copy()
            
            # 複製列寬
            for col_idx, col_dim in new_ws.column_dimensions.items():
                copied_ws.column_dimensions[col_idx].width = col_dim.width
            
            # 複製行高
            for row_idx, row_dim in new_ws.row_dimensions.items():
                copied_ws.row_dimensions[row_idx].height = row_dim.height
        
        # 保存
        merged_wb.save(output_path)
        
        size = os.path.getsize(output_path)
        print(f"[OK] 已生成: {output_path}")
        print(f"  大小: {size:,} bytes ({size / 1024:.1f} KB)")
        
        return output_path
    
    def _clean_sheet_name(self, name: str) -> str:
        """清理工作表名稱（移除非法字符）"""
        illegal = ['\\', '/', '*', '?', ':', '[', ']']
        for ch in illegal:
            name = name.replace(ch, '_')
        return name[:31] if len(name) > 31 else name
    
    def close(self):
        """關閉工作簿"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.worksheet = None
            self.is_loaded = False
