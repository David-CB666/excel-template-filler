#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
統一模板填充器 - 自動選擇引擎

根據模板特徵自動選擇 ZIP 引擎或 openpyxl 引擎

作者: David-CB666
版本: v2.1（融合版）
日期: 2026-06-02
"""

import os
import sys
import io
import json
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple

from engines import create_engine, BaseEngine, ZIPEngine, OpenPYXLEngine


class TemplateFiller:
    """
    統一模板填充器
    
    自動檢測模板類型並選擇最優引擎：
    - 含圖片/打印設置 → ZIP 引擎
    - 無圖片 → openpyxl 引擎
    
    使用示例：
        filler = TemplateFiller("數據源.xlsx", "模板.xlsx")
        placeholders = filler.scan_placeholders()
        output_files = filler.fill_and_export(2, 11, "pdf", "./output")
    """
    
    def __init__(
        self,
        data_source: str = None,
        template: str = None,
        engine_type: str = "auto"
    ):
        """
        初始化模板填充器
        
        Args:
            data_source: 數據源 Excel 路徑
            template: 模板 Excel 路徑
            engine_type: 引擎類型
                "auto" - 自動檢測
                "openpyxl" - 強制使用 openpyxl
                "zip" - 強制使用 ZIP
        """
        self.data_source_path = Path(data_source) if data_source else None
        self.template_path = Path(template) if template else None
        self.engine_type = engine_type
        
        # 引擎實例
        self.engine: Optional[BaseEngine] = None
        
        # 數據
        self.data_list: List[Dict[str, Any]] = []
        self.field_map: Dict[str, str] = {}
        
        # 載入模板
        if self.template_path:
            self._load_template()
    
    def _load_template(self):
        """載入模板"""
        if not self.template_path.exists():
            raise FileNotFoundError(f"模板不存在: {self.template_path}")
        
        # 創建引擎
        self.engine = create_engine(
            template_path=str(self.template_path),
            engine_type=self.engine_type
        )
        
        # 載入模板
        if not self.engine.load_template(str(self.template_path)):
            raise RuntimeError(f"載入模板失敗: {self.template_path}")
        
        # 顯示引擎信息
        engine_name = "ZIP" if isinstance(self.engine, ZIPEngine) else "openpyxl"
        print(f"[INFO] 使用引擎: {engine_name}")
    
    def has_images(self) -> bool:
        """檢測模板是否含圖片"""
        return isinstance(self.engine, ZIPEngine)
    
    def scan_placeholders(self) -> List[str]:
        """
        掃描模板中的佔位符
        
        Returns:
            佔位符列表
        """
        if not self.engine:
            raise RuntimeError("請先載入模板")
        
        return self.engine.scan_placeholders()
    
    def load_data(
        self,
        data_source: str = None,
        field_map: Dict[str, str] = None
    ):
        """
        載入數據源
        
        Args:
            data_source: 數據源 Excel 路徑
            field_map: 字段映射 {佔位符: 字段名}
        """
        if data_source:
            self.data_source_path = Path(data_source)
        
        if field_map:
            self.field_map = field_map
        
        if not self.data_source_path:
            raise RuntimeError("請提供數據源路徑")
        
        if not self.data_source_path.exists():
            raise FileNotFoundError(f"數據源不存在: {self.data_source_path}")
        
        # 使用 openpyxl 讀取數據
        from openpyxl import load_workbook
        
        wb = load_workbook(self.data_source_path, data_only=True)
        ws = wb.active
        
        # 讀取表頭
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        
        # 讀取數據
        self.data_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):  # 跳過空行
                data_row = {}
                for i, header in enumerate(headers):
                    if header and i < len(row):
                        data_row[header] = row[i]
                self.data_list.append(data_row)
        
        wb.close()
        
        print(f"[OK] 已載入數據: {len(self.data_list)} 條記錄")
    
    def validate_data(self) -> Dict[str, Any]:
        """
        驗證數據完整性
        
        Returns:
            驗證結果
        """
        if not self.engine:
            raise RuntimeError("請先載入模板")
        
        # 獲取佔位符
        placeholders = self.scan_placeholders()
        
        # 獲取字段名
        field_names = set(self.field_map.values()) if self.field_map else set(placeholders)
        
        # 獲取數據源字段
        if self.data_list:
            data_fields = set()
            for row in self.data_list:
                data_fields.update(row.keys())
        else:
            data_fields = set()
        
        # 檢查缺失字段
        missing = field_names - data_fields
        
        return {
            "valid": len(missing) == 0,
            "total_rows": len(self.data_list),
            "missing_fields": list(missing),
            "placeholders": placeholders,
            "field_names": list(field_names),
            "data_fields": list(data_fields)
        }
    
    def fill_and_export(
        self,
        start_row: int = None,
        end_row: int = None,
        export_format: str = "excel",
        output_dir: str = "./output",
        field_map: Dict[str, str] = None,
        **kwargs
    ) -> List[str]:
        """
        填充模板並導出
        
        Args:
            start_row: 開始行（數據源，1-based，不包含表頭）
            end_row: 結束行
            export_format: 導出格式
                "excel" - 合併為單個 Excel
                "pdf" - 導出多個 PDF（需要 win32com）
                "both" - 兩者都
            output_dir: 輸出目錄
            field_map: 字段映射 {佔位符: 字段名}
        
        Returns:
            導出的文件路徑列表
        """
        if not self.engine:
            raise RuntimeError("請先載入模板")
        
        if not self.data_list:
            raise RuntimeError("請先載入數據")
        
        if field_map:
            self.field_map = field_map
        
        if not self.field_map:
            raise RuntimeError("請提供字段映射")
        
        # 創建輸出目錄
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        # 選擇數據行
        if start_row is None:
            start_row = 1
        if end_row is None:
            end_row = len(self.data_list)
        
        selected_data = self.data_list[start_row - 1:end_row]
        
        output_files = []
        
        if export_format == "excel":
            # 合併為單個 Excel
            timestamp = self._get_timestamp()
            output_file = output_path / f"Merged_{timestamp}.xlsx"
            
            result = self.engine.fill_and_export(
                data_list=selected_data,
                field_map=self.field_map,
                output_path=str(output_file)
            )
            
            output_files.append(result)
        
        elif export_format == "pdf":
            # 導出多個 PDF（需要 win32com）
            try:
                import win32com.client
                
                # 先生成 Excel
                temp_excel = output_path / "temp.xlsx"
                self.engine.fill_and_export(
                    data_list=selected_data,
                    field_map=self.field_map,
                    output_path=str(temp_excel)
                )
                
                # 導出 PDF
                excel_app = win32com.client.Dispatch("Excel.Application")
                excel_app.Visible = False
                
                wb = excel_app.Workbooks.Open(str(temp_excel.absolute()))
                
                for i, sheet in enumerate(wb.Sheets):
                    pdf_name = selected_data[i].get(
                        list(self.field_map.values())[0],
                        f"Sheet_{i + 1}"
                    )
                    pdf_path = output_path / f"{pdf_name}.pdf"
                    sheet.ExportAsFixedFormat(0, str(pdf_path.absolute()))  # 0 = xlTypePDF
                    output_files.append(str(pdf_path))
                
                wb.Close(False)
                excel_app.Quit()
                
                # 刪除臨時文件
                temp_excel.unlink()
                
                print(f"[OK] 已生成 {len(output_files)} 個 PDF 文件")
                
            except ImportError:
                print("[WARNING] win32com 未安裝，無法導出 PDF")
                print("[INFO] 請安裝: pip install pywin32")
        
        elif export_format == "both":
            # 兩者都導出
            excel_files = self.fill_and_export(
                start_row, end_row, "excel", output_dir, field_map
            )
            pdf_files = self.fill_and_export(
                start_row, end_row, "pdf", output_dir, field_map
            )
            output_files = excel_files + pdf_files
        
        return output_files
    
    def _get_timestamp(self) -> str:
        """獲取時間戳"""
        from datetime import datetime
        return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    
    def close(self):
        """關閉引擎"""
        if self.engine and hasattr(self.engine, 'close'):
            self.engine.close()


def main():
    """命令行入口"""
    import argparse
    
    parser = argparse.ArgumentParser(description="統一模板填充器")
    parser.add_argument("--template", required=True, help="模板 Excel 路徑")
    parser.add_argument("--data", help="數據源 Excel 路徑")
    parser.add_argument("--config", help="配置文件 JSON（包含 data 和 fields）")
    parser.add_argument("--engine", choices=["auto", "openpyxl", "zip"], default="auto", help="引擎類型")
    parser.add_argument("--output", default="./output.xlsx", help="輸出路徑")
    parser.add_argument("--scan", action="store_true", help="僅掃描佔位符")
    
    args = parser.parse_args()
    
    # 創建填充器
    filler = TemplateFiller(
        template=args.template,
        engine_type=args.engine
    )
    
    if args.scan:
        # 僅掃描
        placeholders = filler.scan_placeholders()
        print(f"\n發現佔位符: {placeholders}")
        
        engine_name = "ZIP" if isinstance(filler.engine, ZIPEngine) else "openpyxl"
        print(f"引擎類型: {engine_name}")
    
    elif args.config:
        # 從配置文件讀取
        with open(args.config, 'r', encoding='utf-8') as f:
            cfg = json.load(f)
        
        filler.field_map = cfg.get('fields', {})
        filler.data_list = cfg.get('data', [])
        
        output_path = filler.fill_and_export(
            output_dir=str(Path(args.output).parent),
            field_map=filler.field_map
        )
        
        print(f"\n✅ 完成！請在 Excel 中打開驗證: {output_path}")
    
    elif args.data:
        # 從數據源讀取
        filler.load_data(args.data)
        
        placeholders = filler.scan_placeholders()
        
        # 自動生成字段映射
        field_map = {ph: ph for ph in placeholders}
        
        output_files = filler.fill_and_export(
            field_map=field_map,
            output_dir=str(Path(args.output).parent)
        )
        
        print(f"\n✅ 完成！已生成 {len(output_files)} 個文件")
    
    else:
        print("[ERROR] 請提供 --config 或 --data 參數")
        sys.exit(1)
    
    filler.close()


if __name__ == "__main__":
    main()
